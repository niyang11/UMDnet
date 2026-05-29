import os

# ===================== 设备选择（放在 import torch 前面） =====================
# 可选: "auto" / "cpu" / "cuda"
DEVICE_MODE = "cpu"
GPU_ID = "1"   # 当 DEVICE_MODE="cuda" 或 "auto" 时使用

if DEVICE_MODE == "cpu":
    os.environ["CUDA_VISIBLE_DEVICES"] = "1"
elif DEVICE_MODE in ["cuda", "auto"]:
    os.environ["CUDA_VISIBLE_DEVICES"] = GPU_ID
else:
    raise ValueError(f"不支持的 DEVICE_MODE: {DEVICE_MODE}")

import sys
sys.path.append('/home/project/ny/pythonProject')
sys.path.append('/home/project/ny/pythonProject/improved_diffusion_condition_LD_fre')

import time
import torch
import numpy as np
import scipy.io
import scipy.io as sio
from scipy.io import savemat
from torch.utils.data import Dataset, DataLoader

import pandas as pd
from sklearn.metrics import mean_squared_error, r2_score
from scipy.stats import pearsonr

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ===================== 信号长度（全局使用） =====================
fn = 8192  # 1D NMR 频谱长度

# ===================== 模型导入 =====================
from ISTA_net.model.ISTA_PDF_gau import ISTANet_PDF_GAU
from ISTA_net.model.ISTA_PDF_GMM import ISTANet_PDF_GMM
from Wave_Unet.model.model2 import WaveUNet2
from Wave_Unet.model.model3 import WaveUNet3
from Wave_Unet.model.model0_attention0 import WaveUNet0_attention0
from ISTA_net.utils.dataset import load_phi_qinit
from ISTA_net.model.ISTARB_PdfAMGau import *
from ISTA_net.model.ISTARB_PdfAMGmm import *
from Wave_Unet.model.unet_basic import *
from Wave_Unet.model.model2_3_attention import *
from Wave_Unet.model.unet_basic_att import *
from ISTA_net.model.ISTA_diffusion import *
from Wave_Unet.model.model0_fqy import *
from Wave_Unet.model.model0_uncertain import *
from Wave_Unet.model.model0_Prelu import *
from Wave_Unet.model.model0_uncertain_Prelu import *
from Wave_Unet.model.model2_uncertain import *
from Wave_Unet.model.model20 import WaveUNet20
from ISTA_net.model.ISTA_uncertain import *
from ISTA_net.model.ISTA import ISTANet
from ISTA_net.model.ISTA_ljw import *
from ISTA_net.model.ISTMamba import ISTMamba
from DNunet.model_dn import *
from DNunet.unet_basic_dn import *
from ISTA_net.model.ISTA_ljw_uncertain import *
from DNunet.unet_basic_dn_uncertain import *
from ISTA_net.model.ISTA_attention0 import ISTANet_attention0
from Wave_Unet.model.waveumamba import WaveUMamba
from Wave_Unet.model.waveumamba_flex import WaveUMambaFlex
from ISTA_net.model.ISTA_ljw_evident import *
from Wave_Unet.model.WaveUGateMamba import WaveUGateMamba
# ================================================

# ===================== 根据 DEVICE_MODE 选择 device =====================
if DEVICE_MODE == "cpu":
    device = torch.device("cpu")

elif DEVICE_MODE == "cuda":
    if not torch.cuda.is_available():
        raise RuntimeError("你指定了使用 GPU，但当前环境没有可用 CUDA 设备。")
    device = torch.device("cuda")

elif DEVICE_MODE == "auto":
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

print(f"当前测试设备: {device}")
if device.type == "cuda":
    print(f"当前使用 GPU_ID: {GPU_ID}")


# ===================== 数据集类定义 =====================
class CustomDataset1D(Dataset):
    def __init__(self, fftn_data, fft_data):
        self.fftn_data = torch.tensor(fftn_data, dtype=torch.float32)
        self.fft_data = torch.tensor(fft_data, dtype=torch.float32)

    def __len__(self):
        return self.fftn_data.shape[0]

    def __getitem__(self, idx):
        return self.fftn_data[idx], self.fft_data[idx]


class CustomDataset1DISTA(Dataset):
    def __init__(self, fftn_data, fft_data):
        # (N, 8192, 1) -> (N, 8192)
        self.fftn_data = torch.tensor(fftn_data, dtype=torch.float32).squeeze(2)
        self.fft_data = torch.tensor(fft_data, dtype=torch.float32).squeeze(2)

    def __len__(self):
        return self.fftn_data.shape[0]

    def __getitem__(self, idx):
        return self.fftn_data[idx], self.fft_data[idx]


class CustomDataset2D(Dataset):
    def __init__(self, fftn_data, fft_data):
        self.fftn_data = fftn_data
        self.fft_data = fft_data

    def __len__(self):
        return self.fftn_data.shape[0]

    def __getitem__(self, idx):
        fftn_sample = self.fftn_data[idx].reshape(-1, 1)
        fft_sample = self.fft_data[idx].reshape(-1, 1)
        return fftn_sample, fft_sample


class CustomDataset3D(Dataset):
    def __init__(self, fftn_data, for_ista=False):
        self.fftn_data = torch.tensor(fftn_data, dtype=torch.float32)
        self.for_ista = for_ista

    def __len__(self):
        return self.fftn_data.shape[0]

    def __getitem__(self, idx):
        fftn_sample = self.fftn_data[idx]
        fftn_sample = fftn_sample.reshape(-1, 1)
        return fftn_sample


# ===================== 数据加载函数 =====================
def load_1d_data(file_path, fftn_key, fft_key=None, for_ista=False):
    mat_data = scipy.io.loadmat(file_path)
    fftn_data = mat_data[fftn_key]
    fft_data = mat_data[fft_key] if fft_key else fftn_data

    fftn_data = fftn_data.transpose(1, 0)
    fftn_data = fftn_data[np.newaxis, :, :].astype(np.float32)
    fft_data = fft_data.transpose(1, 0)[np.newaxis, :, :].astype(np.float32) if fft_key else fftn_data

    min_len = min(fftn_data.shape[2], fft_data.shape[2])
    fftn_data = fftn_data[:, :, :min_len]
    fft_data = fft_data[:, :, :min_len]

    batch_size = 4
    dataset = CustomDataset1DISTA(fftn_data, fft_data) if for_ista else CustomDataset1D(fftn_data, fft_data)
    loader = DataLoader(dataset, batch_size=batch_size, shuffle=False)
    return loader


def load_mix_moise(sequence_number, for_ista=False):
    file_path = '/home/project/ny/matlab/wave_unet/mix_noise.mat'
    fftn_key = f'real_FFTN{sequence_number}'
    fft_key = 'real_FFT'
    return load_1d_data(file_path, fftn_key, fft_key, for_ista=for_ista)


def load_aq_moise(sequence_number, for_ista=False):
    file_path = '/home/project/ny/matlab/wave_unet/aq_noise.mat'
    fftn_key = f'real_FFTN{sequence_number}'
    fft_key = 'real_FFT'
    return load_1d_data(file_path, fftn_key, fft_key, for_ista=for_ista)


def load_danguchun_moise(sequence_number, for_ista=False):
    file_path = '/home/project/ny/matlab/wave_unet/danguchun_noise.mat'
    fftn_key = f'real_FFTN{sequence_number}'
    fft_key = 'real_FFT'
    return load_1d_data(file_path, fftn_key, fft_key, for_ista=for_ista)


def load_2d_danbai_0902(number, for_ista=False):
    mat_data = scipy.io.loadmat('/home/project/ny/matlab/wave_unet/2d_data/test_2d_danbai_0902.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=512, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=512, shuffle=False)
    return loader


def load_2d_gb1_0902(number, for_ista=False):
    mat_data = scipy.io.loadmat('/home/project/ny/matlab/wave_unet/2d_data/test_2d_gb1_0902.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=512, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=512, shuffle=False)
    return loader


def load_2d_titr31(number, for_ista=False):
    mat_data = sio.loadmat('/home/project/ny/matlab/wave_unet/2d_data/test_2d_titr31.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    return loader


def load_hnco_xy(number, for_ista=False):
    mat_data = scipy.io.loadmat('/home/project/ny/matlab/wave_unet/3d_data/test_3d_dataxy.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    return loader


def load_hnco_xz(number, for_ista=False):
    mat_data = scipy.io.loadmat('/home/project/ny/matlab/wave_unet/3d_data/test_3d_dataxz.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    return loader


def load_hnco_yz(number, for_ista=False):
    mat_data = scipy.io.loadmat('/home/project/ny/matlab/wave_unet/3d_data/test_3d_datayz.mat')
    data_keys = {0: 'real_FFTN0', 1: 'real_FFTN1', 2: 'real_FFTN2', 3: 'real_FFTN3'}
    label_key = 'real_FFT'
    fftn_data = mat_data[data_keys[number]].astype(np.float32)
    fft_data = mat_data[label_key].astype(np.float32)

    if for_ista:
        dataset = torch.utils.data.TensorDataset(
            torch.tensor(fftn_data, dtype=torch.float32),
            torch.tensor(fft_data, dtype=torch.float32)
        )
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    else:
        dataset = CustomDataset2D(fftn_data, fft_data)
        loader = DataLoader(dataset, batch_size=256, shuffle=False)
    return loader


# ===================== 指标工具函数 =====================
def compute_reconstruction_residual(den, true):
    return den - true


def get_snr(prediction, freq1, freq2):
    prediction = np.asarray(prediction).reshape(-1)
    signal_region = prediction[freq1:freq2]
    signal_std = np.std(signal_region)
    prediction_max = np.max(prediction)
    snr = prediction_max / (signal_std + 1e-12)
    return snr


def get_model_stats(model, model_path):
    try:
        model_size_mb = os.path.getsize(model_path) / (1024 * 1024)
    except OSError:
        model_size_mb = float("nan")

    total_params = sum(p.numel() for p in model.parameters())
    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    return model_size_mb, int(total_params), int(trainable_params)


# ===================== 模型类型判定 =====================
def is_ist_model(model_class):
    name = model_class.__name__
    return ('IST' in name) or ('ISTA' in name)


def build_model_instance(model_class, n_layers, kwargs=None):
    if kwargs is None:
        kwargs = {}

    if is_ist_model(model_class):
        try:
            return model_class(LayerNo=n_layers, **kwargs)
        except TypeError:
            pass
        try:
            return model_class(n_layers=n_layers, **kwargs)
        except TypeError:
            pass
        try:
            return model_class(**kwargs)
        except TypeError:
            raise RuntimeError(f"{model_class.__name__} 实例化失败，请检查构造参数。")

    try:
        return model_class(n_layers=n_layers, channels_interval=24, **kwargs)
    except TypeError:
        pass
    try:
        return model_class(n_layers=n_layers, **kwargs)
    except TypeError:
        pass
    try:
        return model_class(**kwargs)
    except TypeError:
        raise RuntimeError(f"{model_class.__name__} 实例化失败，请检查构造参数。")


# ===================== 测试函数 =====================
def test_model(
    model,
    model_path,
    test_loader,
    device,
    save_path,
    model_number,
    input_type,
    variable_name="result",
    Phi_ideal=None,
    Qinit_ideal=None,
    display_name=None,
    test_name=None,
):
    model = model.to(device)

    state = torch.load(model_path, map_location=device)
    if isinstance(state, dict) and "state_dict" in state:
        state = state["state_dict"]

    new_state = {}
    for k, v in state.items():
        nk = k[7:] if k.startswith("module.") else k
        new_state[nk] = v

    model.load_state_dict(new_state, strict=False)

    model_size_mb, total_params, trainable_params = get_model_stats(model, model_path)

    model.eval()
    test_results = []
    inference_times = []
    all_labels = []
    all_fftn = []

    r2 = float("nan")
    pearson_r2 = float("nan")
    rmsd = float("nan")
    snr_increment = float("nan")

    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    with torch.no_grad():
        for data in test_loader:
            if isinstance(data, (list, tuple)):
                fftn_data, labels_data = data
                labels_data = labels_data.to(device).float()
            else:
                fftn_data = data
                labels_data = None

            fftn_data = fftn_data.to(device).float()

            if fftn_data.ndimension() >= 2 and fftn_data.shape[1] == 32768:
                fftn_data = fftn_data.view(4, fn, 1)

            if device.type == "cuda":
                start_event = torch.cuda.Event(enable_timing=True)
                end_event = torch.cuda.Event(enable_timing=True)

            # ================= ISTARB 系列 =================
            if 'ISTARB' in model.__class__.__name__:
                if len(fftn_data.shape) == 3:
                    fftn_data = fftn_data.squeeze(2)

                if device.type == "cuda":
                    torch.cuda.empty_cache()
                    _ = model(torch.randn(4, fn).to(device), Phi_ideal)
                    start_event.record()
                    output = model(fftn_data, Phi_ideal)
                    end_event.record()
                    torch.cuda.synchronize()
                    inference_time_ms = start_event.elapsed_time(end_event)
                else:
                    t0 = time.time()
                    output = model(fftn_data, Phi_ideal)
                    t1 = time.time()
                    inference_time_ms = (t1 - t0) * 1000.0

            # ================= IST / ISTA / ISTMamba / ISTNet =================
            elif Phi_ideal is not None and Qinit_ideal is not None:
                if len(fftn_data.shape) == 3:
                    fftn_data = fftn_data.squeeze(2)

                if device.type == "cuda":
                    torch.cuda.empty_cache()
                    _ = model(torch.randn(4, fn).to(device), Phi_ideal, Qinit_ideal)
                    start_event.record()
                    model_output = model(fftn_data, Phi_ideal, Qinit_ideal)
                    end_event.record()
                    torch.cuda.synchronize()
                    inference_time_ms = start_event.elapsed_time(end_event)
                else:
                    t0 = time.time()
                    model_output = model(fftn_data, Phi_ideal, Qinit_ideal)
                    t1 = time.time()
                    inference_time_ms = (t1 - t0) * 1000.0

                if isinstance(model_output, (list, tuple)):
                    output = model_output[0]
                else:
                    output = model_output

            # ================= 只用 Phi 的其它 IST 变体 =================
            elif Phi_ideal is not None:
                if len(fftn_data.shape) == 3:
                    fftn_data = fftn_data.squeeze(2)

                if device.type == "cuda":
                    torch.cuda.empty_cache()
                    _ = model(torch.randn(4, fn).to(device), Phi_ideal)
                    start_event.record()
                    model_output = model(fftn_data, Phi_ideal)
                    end_event.record()
                    torch.cuda.synchronize()
                    inference_time_ms = start_event.elapsed_time(end_event)
                else:
                    t0 = time.time()
                    model_output = model(fftn_data, Phi_ideal)
                    t1 = time.time()
                    inference_time_ms = (t1 - t0) * 1000.0

                if isinstance(model_output, (list, tuple)):
                    output = model_output[0]
                else:
                    output = model_output

            # ================= WaveUNet / WaveUMamba 等 =================
            else:
                if input_type == 'normal' and model_number < 10:
                    if fftn_data.ndimension() == 2:
                        pass
                    elif fftn_data.shape[1] != fn or fftn_data.shape[2] != 1:
                        fftn_data = fftn_data.permute(0, 2, 1)
                elif input_type == 'transpose':
                    if fftn_data.shape[1] != 1 or fftn_data.shape[2] != fn:
                        fftn_data = fftn_data.permute(0, 2, 1)

                fftn_data = fftn_data.permute(0, 2, 1)

                bs = fftn_data.shape[0]
                if device.type == "cuda":
                    torch.cuda.empty_cache()
                    _ = model(torch.randn(bs, 1, fn).to(device))
                    start_event.record()
                    output = model(fftn_data)
                    end_event.record()
                    torch.cuda.synchronize()
                    inference_time_ms = start_event.elapsed_time(end_event)
                else:
                    t0 = time.time()
                    output = model(fftn_data)
                    t1 = time.time()
                    inference_time_ms = (t1 - t0) * 1000.0

                if isinstance(output, tuple):
                    output = output[0]

            inference_times.append(inference_time_ms)

            if hasattr(output, "shape") and output.shape[0] == 4:
                output = output.reshape(1, 4 * fn)

            output = output.squeeze()
            test_results.append(output.detach().cpu().numpy())

            if labels_data is not None:
                all_labels.append(labels_data.detach().cpu().numpy())
            all_fftn.append(fftn_data.detach().cpu().numpy())

    predictions = np.vstack(test_results)
    maxv = np.max(predictions) if np.max(predictions) != 0 else 1.0
    predictions = predictions / maxv
    predictions = np.squeeze(predictions)

    labels_all = np.vstack(all_labels) if len(all_labels) > 0 else None
    fftn_all = np.vstack(all_fftn)

    savemat(save_path, {variable_name: predictions})
    print(f"Predictions saved to {save_path}")

    avg_time = float(np.mean(inference_times)) if len(inference_times) > 0 else float("nan")

    if labels_all is not None:
        labels_all = np.squeeze(labels_all)

        fftn_all_ = fftn_all.reshape(-1, 1)
        predictions_ = predictions.reshape(-1, 1)
        labels_all_ = labels_all.reshape(-1, 1)

        _ = compute_reconstruction_residual(labels_all_, predictions_)

        snr_input = get_snr(fftn_all_, -150, -100)
        snr_output = get_snr(predictions_, -150, -100)
        snr_increment = snr_output / (snr_input + 1e-12)

        rmsd = np.sqrt(mean_squared_error(labels_all_, predictions_))
        r2 = r2_score(labels_all_, predictions_)
        pearson_r, _ = pearsonr(labels_all_.ravel(), predictions_.ravel())
        pearson_r2 = float(pearson_r ** 2)

    tag = f"[{display_name} | {test_name}]"
    print(f"{tag} 模型大小(MB): {model_size_mb:.4f}")
    print(f"{tag} 参数量: {total_params}")
    print(f"{tag} 平均推理时间(ms): {avg_time:.4f}")
    print(f"{tag} R²: {r2:.4f}")
    print(f"{tag} Pearson r²: {pearson_r2:.4f}")
    print(f"{tag} RMSD: {rmsd:.4f}")
    print(f"{tag} KSNR: {snr_increment:.4f}")

    metrics = {
        '模型大小(MB)': float(model_size_mb),
        '参数量': int(total_params),
        '平均推理时间(ms)': avg_time,
        'R²': float(r2) if np.isfinite(r2) else r2,
        'Pearson r²': float(pearson_r2) if np.isfinite(pearson_r2) else pearson_r2,
        'RMSD': float(rmsd) if np.isfinite(rmsd) else rmsd,
        'KSNR': float(snr_increment) if np.isfinite(snr_increment) else snr_increment,
    }
    return predictions, metrics


# ===================== 模型列表 =====================
ABLA_ROOT = "/home/project/ny/pythonProject/Wave_Unet/best_model/ablation_M1M2_k2k4k12/"

models = [
    ("WaveUNet0", WaveUNet0,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/best_model0_20000SS/best_model0_20000SS_250330.pth",
     12, 37, 'normal', {}),

    ("1041", WaveUMambaFlex,
     os.path.join(ABLA_ROOT, "M1_k4_R1", "best.pth"),
     12, 1041, 'normal',
     {
         "mamba_from": 4,
         "mamba_version": 1,
         "mamba_internal_residual": True,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "headdim": 8,
         "mamba_dropout": 0.1,
         "out_residual": True,
         "out_activation": "none",
         "gamma_init": 0.1,
     }),

    ("2040", WaveUMambaFlex,
     os.path.join(ABLA_ROOT, "M2_k4_R0", "best.pth"),
     12, 2040, 'normal',
     {
         "mamba_from": 4,
         "mamba_version": 2,
         "mamba_internal_residual": False,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "headdim": 8,
         "mamba_dropout": 0.1,
         "out_residual": True,
         "out_activation": "none",
         "gamma_init": 0.1,
     }),

    ("ISTNet", ISTNet,
     "/home/project/ny/pythonProject/ISTA_net/best_model/ljw_20000SS_layer9/model_dir/net_params_100_8192.pth",
     9, 9, 'normal', {}),

    ("WaveUMamba_m0", WaveUMamba,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/best_waveumamba_20000SS_1222/best_waveumamba_20000SS_251222.pth",
     12, 0, 'normal',
     {
         "mamba_from": 0,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "residual": True,
         "out_activation": "none",
     }),

    ("WaveUMamba_m4", WaveUMamba,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/best_waveumamba_20000SS/best_waveumamba_20000SS_251208.pth",
     12, 38, 'normal',
     {
         "mamba_from": 4,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "residual": True,
         "out_activation": "none",
     }),

    ("WaveUMamba_m8", WaveUMamba,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/fusion_sweep_0126/scheme08_mfrom08/best_scheme08_mfrom08.pth",
     12, 8, 'normal',
     {
         "mamba_from": 8,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "residual": True,
         "out_activation": "none",
     }),

    ("WaveUMamba_m12", WaveUMamba,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/fusion_sweep_0126/scheme12_mfrom12/best_scheme12_mfrom12.pth",
     12, 12, 'normal',
     {
         "mamba_from": 12,
         "d_state": 64,
         "d_conv": 4,
         "expand": 2,
         "residual": True,
         "out_activation": "none",
     }),

    ("Modeldnwave", Modeldnwave,
     "/home/project/ny/pythonProject/Wave_Unet/best_model/best_model_dnunet_20000SS/best_model_DN_20000SS_250331.pth",
     12, 14, 'normal', {}),
]

for model_name, _, ckpt, _, _, _, _ in models:
    if not os.path.exists(ckpt):
        raise FileNotFoundError(f"{model_name} 的权重文件不存在: {ckpt}")


# ===================== 普通模型测试集 =====================
test_loaders_info = [
    ("mix_4", load_mix_moise(4, for_ista=False)),
    ("mix_1", load_mix_moise(1, for_ista=False)),
    ("mix_2", load_mix_moise(2, for_ista=False)),
    ("mix_3", load_mix_moise(3, for_ista=False)),


    ("dan_4", load_danguchun_moise(4, for_ista=False)),
    ("dan_1", load_danguchun_moise(1, for_ista=False)),
    ("dan_2", load_danguchun_moise(2, for_ista=False)),
    ("dan_3", load_danguchun_moise(3, for_ista=False)),


    ("aq_4", load_aq_moise(4, for_ista=False)),
    ("aq_1", load_aq_moise(1, for_ista=False)),
    ("aq_2", load_aq_moise(2, for_ista=False)),
    ("aq_3", load_aq_moise(3, for_ista=False)),

    #
    # ("titr31_1", load_2d_titr31(1, for_ista=False)),
    # ("titr31_2", load_2d_titr31(2, for_ista=False)),
    # ("titr31_3", load_2d_titr31(3, for_ista=False)),
    #
    # ("danbai_1_0902", load_2d_danbai_0902(1, for_ista=False)),
    # ("danbai_2_0902", load_2d_danbai_0902(2, for_ista=False)),
    # ("danbai_3_0902", load_2d_danbai_0902(3, for_ista=False)),
    #
    # ("gb_1_0902", load_2d_gb1_0902(1, for_ista=False)),
    # ("gb_2_0902", load_2d_gb1_0902(2, for_ista=False)),
    # ("gb_3_0902", load_2d_gb1_0902(3, for_ista=False)),
    #
    # ("HNCO_XY_1", load_hnco_xy(1, for_ista=False)),
    # ("HNCO_XZ_1", load_hnco_xz(2, for_ista=False)),
    # ("HNCO_YZ_1", load_hnco_yz(3, for_ista=False)),
    #
    # ("HNCO_XY_2", load_hnco_xy(1, for_ista=False)),
    # ("HNCO_XZ_2", load_hnco_xz(2, for_ista=False)),
    # ("HNCO_YZ_2", load_hnco_yz(3, for_ista=False)),
    #
    # ("HNCO_XY_3", load_hnco_xy(1, for_ista=False)),
    # ("HNCO_XZ_3", load_hnco_xz(2, for_ista=False)),
    # ("HNCO_YZ_3", load_hnco_yz(3, for_ista=False)),
]

# ===================== IST 模型测试集 =====================
test_loaders_ista_info = [

    ("mix_4", load_mix_moise(4, for_ista=True)),
    ("mix_1", load_mix_moise(1, for_ista=True)),
    ("mix_2", load_mix_moise(2, for_ista=True)),
    ("mix_3", load_mix_moise(3, for_ista=True)),

    ("dan_4", load_danguchun_moise(4, for_ista=True)),
    ("dan_1", load_danguchun_moise(1, for_ista=True)),
    ("dan_2", load_danguchun_moise(2, for_ista=True)),
    ("dan_3", load_danguchun_moise(3, for_ista=True)),

    ("aq_4", load_aq_moise(4, for_ista=True)),
    ("aq_1", load_aq_moise(1, for_ista=True)),
    ("aq_2", load_aq_moise(2, for_ista=True)),
    ("aq_3", load_aq_moise(3, for_ista=True)),


    # ("titr31_1", load_2d_titr31(1, for_ista=True)),
    # ("titr31_2", load_2d_titr31(2, for_ista=True)),
    # ("titr31_3", load_2d_titr31(3, for_ista=True)),
    #
    # ("danbai_1_0902", load_2d_danbai_0902(1, for_ista=True)),
    # ("danbai_2_0902", load_2d_danbai_0902(2, for_ista=True)),
    # ("danbai_3_0902", load_2d_danbai_0902(3, for_ista=True)),
    #
    # ("gb_1_0902", load_2d_gb1_0902(1, for_ista=True)),
    # ("gb_2_0902", load_2d_gb1_0902(2, for_ista=True)),
    # ("gb_3_0902", load_2d_gb1_0902(2, for_ista=True)),  # 按你参考代码保留
    #
    # ("HNCO_XY_1", load_hnco_xy(1, for_ista=True)),
    # ("HNCO_XZ_1", load_hnco_xz(2, for_ista=True)),
    # ("HNCO_YZ_1", load_hnco_yz(3, for_ista=True)),
    #
    # ("HNCO_XY_2", load_hnco_xy(1, for_ista=True)),
    # ("HNCO_XZ_2", load_hnco_xz(2, for_ista=True)),
    # ("HNCO_YZ_2", load_hnco_yz(3, for_ista=True)),
    #
    # ("HNCO_XY_3", load_hnco_xy(1, for_ista=True)),
    # ("HNCO_XZ_3", load_hnco_xz(2, for_ista=True)),
    # ("HNCO_YZ_3", load_hnco_yz(3, for_ista=True)),
]

# ===================== 加载 Phi / Qinit =====================
Phi_ideal_path = r'/home/project/ny/matlab/istmatrix.mat'
Qinit_ideal_path = r'/home/project/ny/matlab/data/Qinit_20000SS'
mat_data_path = r'/home/SSD2_4T/fangqy/train_data/IST/IST_20000_SS.mat'

Phi_input, Qinit = load_phi_qinit(Phi_ideal_path, Qinit_ideal_path, mat_data_path, is_mat=True)
Phi_ideal = torch.tensor(Phi_input, dtype=torch.float32).to(device)
Qinit_ideal = torch.tensor(Qinit, dtype=torch.float32).to(device)

# ===================== 主循环 =====================
all_metrics = {}

for model_info in models:
    display_name, model_class, model_path, n_layers, model_number, input_type, kwargs = model_info

    print("\n" + "=" * 100)
    print(f"开始测试模型: {display_name}")
    print(f"权重路径: {model_path}")
    print("=" * 100)

    if is_ist_model(model_class):
        current_loaders = test_loaders_ista_info
        current_phi = Phi_ideal
        current_qinit = Qinit_ideal
    else:
        current_loaders = test_loaders_info
        current_phi = None
        current_qinit = None

    for test_name, test_loader in current_loaders:
        variable_name = "result"
        save_path = f'/home/project/ny/pythonProject/Wave_Unet/predictions/predictions_all/{display_name}_{test_name}_predictions_{model_number}.mat'
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        model_instance = build_model_instance(model_class, n_layers, kwargs)

        _, metrics = test_model(
            model_instance,
            model_path,
            test_loader,
            device,
            save_path,
            model_number,
            input_type,
            variable_name=variable_name,
            Phi_ideal=current_phi,
            Qinit_ideal=current_qinit,
            display_name=display_name,
            test_name=test_name
        )

        if test_name not in all_metrics:
            all_metrics[test_name] = []

        all_metrics[test_name].append({
            '模型': f"{display_name}_{model_number}",
            '模型大小(MB)': metrics['模型大小(MB)'],
            '参数量': metrics['参数量'],
            '平均推理时间(ms)': metrics['平均推理时间(ms)'],
            'R²': metrics['R²'],
            'Pearson r²': metrics['Pearson r²'],
            'RMSD': metrics['RMSD'],
            'KSNR': metrics['KSNR'],
        })

# ===================== 写入 Excel =====================
excel_save_path = 'metrics_all_models0401.xlsx'

with pd.ExcelWriter(excel_save_path, engine='openpyxl') as writer:
    for sheet_name, metrics_list in all_metrics.items():
        df = pd.DataFrame(metrics_list)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"指标已保存到 Excel 文件：{excel_save_path}")

# ===================== Excel 格式设置 =====================
wb = load_workbook(excel_save_path)
thin_side = Side(border_style="thin", color="000000")
cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 25
        for cell in col:
            cell.font = Font(name='宋体', size=16)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save(excel_save_path)
print(f"Excel 文件格式设置完成，路径：{excel_save_path}")